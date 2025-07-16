from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
import ctypes

from plyer import notification

#CONSTANTE ALERTA
MB_SYSTEMMODAL = 0x00001000

#FUNCIONES
def obtenerDatosXSL():
    try:
        # Abre el archivo excel
        wb = load_workbook('Datos bot.xlsx')
        # Selecciona la hoja
        ws = wb['Hoja1']
        # Obtiene los datos
        email =ws['A2'].value
        pwd =ws['B2'].value
        max_row = ws.max_row        

        lista_temas = []
        for i in range(2, max_row):
            lista_temas.append(ws.cell(row=i, column=3).value)
        print("Estos son los temas a buscar: " + lista_temas.__str__() + "\n")
        print("Esta consola se cerrara luego de ejecutar el bot")
        return email, pwd, lista_temas
    except:
            ctypes.windll.user32.MessageBoxW(0, "El bot fallo al cargar los datos del excel", "EXCEL ERROR", MB_SYSTEMMODAL)

def openWeb(driver, email, pwd, listaTematicas):

   # try:
        driver.get('https://www.bing.com/search?q=Bing+AI&qs=HS&sc=11-0&cvid=AD62DD36244A4856B319047D747E7840&FORM=QBLH&sp=1&lq=0')
        driver.implicitly_wait(5)
        time.sleep(5)

        # Busca el boton de iniciar sesion
        try:
            botonIniciarSesion = driver.find_element(By.XPATH, '/html/body/header/div/a[2]/div[1]/span/input')
            botonIniciarSesion.click()
        except:
            try:
                botonIniciarSesion2 = driver.find_element(By.XPATH, '/html/body/header/div/a[1]/div[1]/span/input')
                botonIniciarSesion2.click()
            except:
                print("No se pudo encontrar el boton de iniciar sesion, revisa el xpath")
        time.sleep(3)

        #Completo user
        try:
            username_input = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div/div/div/div/form/div[2]/div/div[2]/div[1]/div/span/input')
        except:
            try:            
                username_input = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/div/div/form/div[2]/div/div/input')
            except:
                print("No se pudo encontrar el campo de usuario, revisa el xpath")
        username_input.send_keys(email)

        try:
            sig_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div/div/div/div/form/div[2]/div/div[3]/button')
        except:
             sig_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div/div/div/form/div[4]/div/div/div/div/button')
        sig_btn.click()
        time.sleep(5)

         #Completo pass 
        try:         
            pwd_input = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div/div/div/div/form/div[2]/div/div[4]/div[1]/div/span/input')
        except:
            try:
                pwd_input = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div/div/div/div/form/div[2]/div/div[4]/div[1]/div/span/input')
            except:
                # OTRAS FORMAS DE INICIAR SESION
                try:
                    otras_formas_btn = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div/div/div/div/form/div[2]/div/span[1]/div/span')
                    otras_formas_btn.click()
                    time.sleep(3)
                
                    #USAR PWD
                    pwd_option = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div/div/div/div/div[2]/div/div[3]/div[3]/div')
                    pwd_option.click()
                    time.sleep(3)
                except:
                    print("No se pudo encontrar el boton de otras formas de iniciar sesion, continuando...")

                try:         
                    pwd_input = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div[1]/div/div/div/div/div[2]/div[2]/div/form/div[3]/div/div/input')
                except:
                    try:
                        pwd_input = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div/div/div/div/form/div[2]/div/div[4]/div[1]/div/span/input')
                    except:
                        print("No se pudo encontrar el campo de contraseña, revisa el xpath")
                        return 0
        pwd_input.send_keys(pwd)
       

        #Siguiente boton
        try:
            sig_btn2 = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div/div/div/div/form/div[2]/div/div[5]/button')
            sig_btn2.click()
            time.sleep(2)
        except:
            print("No se pudo encontrar el boton de siguiente, continuando...")

        #Mantener sesion iniciada
        try:
            mantener_sesion_btn = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div/div[1]/div/div/div/div/form/div[2]/div/div[5]/button[2]')
            mantener_sesion_btn.click()
        except:
            try:
                mantener_sesion_btn = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div/div[1]/div/div/div/div/form/div[1]/button/span/svg[1]')
                mantener_sesion_btn.click()
            except:
                print("No se pudo encontrar el boton de mantener sesion iniciada, continuando...")

        time.sleep(2)
        
        #Inserto dato en el buscador
        txtBoxSearch = driver.find_element(By.XPATH, '/html/body/header/form/div/input[1]')
        txtBoxSearch.clear()
        txtBoxSearch.send_keys('clima hoy')
        txtBoxSearch.submit()
        time.sleep(5)

        # Ciclo de busqueda
        for i in range(0, 30):
            # Busca una tematica de una lista de tematicas
            buscador = driver.find_element(By.XPATH, '/html/body/header/form/div/input[1]')
            buscador.clear()
            buscador.send_keys(listaTematicas[i])
            buscador.submit()
            time.sleep(10)


    # except:
	#     ctypes.windll.user32.MessageBoxW(0, "El bot fallo al ejecutar", "BOT ERROR", MB_SYSTEMMODAL)
         
def driver_init():
    chrome_options = webdriver.ChromeOptions()

    # Configuracion
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument('log-level=3')
    chrome_options.add_argument('--lang=es')

    # Inicializa el driver con las opciones especificadas
    driver = webdriver.Chrome(chrome_options)
    return driver

def main():
    email, pwd, lista_temas = obtenerDatosXSL()
    # Carga el driver
    driver = driver_init()
    openWeb(driver, email, pwd, lista_temas)
main()
